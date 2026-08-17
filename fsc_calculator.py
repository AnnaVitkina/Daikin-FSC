from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from config import (
    CALCULATION_BASIS_DIR,
    OUTPUT_DIR,
    PROCESSING_DIR,
    WORKSPACE_ROOT,
)

PROCESSING_SHEET_WITH_TAXES = "Prices with taxes"
PROCESSING_SHEET_WO_TAXES = "Prices wo taxes"
ROTALIS_PT_SHEET = "Rotalis PT"

FSC_MULTIPLIER = 0.25
WO_TAXES_DECIMALS = 3
FSC_PERCENT_DECIMALS = 2

OUTPUT_COLUMNS = ("Date", "Country", "Diesel price", "FSC value, %")

HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_FONT = Font(bold=True, color="1F4E78")
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
COLUMN_WIDTHS = {
    "Date": 14,
    "Country": 10,
    "Diesel price": 14,
    "FSC value, %": 14,
}


def list_processed_files(processing_dir: Path | None = None) -> list[Path]:
    """Return available processed xlsx files in the processing folder."""
    folder = processing_dir or PROCESSING_DIR
    if not folder.exists():
        raise FileNotFoundError(f"Processing folder not found: {folder}")

    files = sorted(
        path for path in folder.glob("*.xlsx")
        if not path.name.startswith("~$")
    )
    if not files:
        raise FileNotFoundError(f"No processed .xlsx files found in: {folder}")

    return files


def choose_processed_file(processing_dir: Path | None = None) -> Path:
    """Prompt the user to choose a processed file from the processing folder."""
    files = list_processed_files(processing_dir)

    print("Available processed files:")
    for index, file_path in enumerate(files, start=1):
        print(f"  {index}. {file_path.name}")

    while True:
        choice = input("Enter file number: ").strip()
        if not choice.isdigit():
            print("Please enter a valid number.")
            continue

        selected_index = int(choice)
        if 1 <= selected_index <= len(files):
            return files[selected_index - 1]

        print(f"Please enter a number between 1 and {len(files)}.")


def find_calculation_basis_file(basis_dir: Path | None = None) -> Path:
    """Locate the Calculation basis xlsx file."""
    folder = basis_dir or CALCULATION_BASIS_DIR
    direct_path = folder / "Calculation basis.xlsx"
    if direct_path.exists():
        return direct_path

    matches = sorted((WORKSPACE_ROOT / "input" / "rules").rglob("Calculation basis.xlsx"))
    matches = [path for path in matches if not path.name.startswith("~$")]
    if not matches:
        raise FileNotFoundError(
            "Calculation basis.xlsx was not found under input/rules."
        )

    return matches[0]


def get_non_rotalis_sheet(sheet_names: list[str]) -> str:
    """Return the calculation basis tab that is not Rotalis PT."""
    for sheet_name in sheet_names:
        if sheet_name.strip().casefold() != ROTALIS_PT_SHEET.casefold():
            return sheet_name

    raise ValueError(
        f"Could not find a calculation basis tab other than '{ROTALIS_PT_SHEET}'."
    )


def load_fuel_costs(basis_path: Path, sheet_name: str) -> dict[str, float]:
    """Load country fuel cost values from a calculation basis tab."""
    basis_df = pd.read_excel(basis_path, sheet_name=sheet_name, engine="openpyxl")
    if basis_df.empty:
        raise ValueError(f"Calculation basis tab '{sheet_name}' is empty.")

    fuel_costs: dict[str, float] = {}
    fuel_row = basis_df.iloc[0]

    for column_name in basis_df.columns:
        if str(column_name).strip().casefold() == "country":
            continue

        value = fuel_row[column_name]
        if pd.isna(value):
            continue

        country_code = str(column_name).strip().upper()
        fuel_costs[country_code] = float(value)

    if not fuel_costs:
        raise ValueError(
            f"No fuel cost values found in calculation basis tab '{sheet_name}'."
        )

    return fuel_costs


def round_numeric(value: float, decimals: int) -> float:
    return round(float(value), decimals)


def calculate_fsc_percent_with_taxes(
    diesel_price: float,
    fuel_cost: float,
) -> float:
    return round_numeric(
        ((diesel_price / fuel_cost) - 1) * FSC_MULTIPLIER * 100,
        FSC_PERCENT_DECIMALS,
    )


def calculate_fsc_percent_wo_taxes(
    diesel_price: float,
    fuel_cost: float,
) -> float:
    rounded_price = round_numeric(diesel_price, WO_TAXES_DECIMALS)
    rounded_fuel_cost = round_numeric(fuel_cost, WO_TAXES_DECIMALS)
    return round_numeric(
        ((rounded_price / rounded_fuel_cost) - 1) * FSC_MULTIPLIER * 100,
        FSC_PERCENT_DECIMALS,
    )


def build_fsc_dataframe(
    prices_df: pd.DataFrame,
    fuel_costs: dict[str, float],
    *,
    wo_taxes: bool,
) -> pd.DataFrame:
    """Build the FSC output table for one processed price tab."""
    records: list[dict[str, object]] = []

    for row in prices_df.itertuples(index=False):
        country = str(row.consumer).strip().upper()
        fuel_cost = fuel_costs.get(country)
        if fuel_cost is None:
            continue

        diesel_price = float(row.price_per_unit)
        if wo_taxes:
            fsc_value = calculate_fsc_percent_wo_taxes(diesel_price, fuel_cost)
        else:
            fsc_value = calculate_fsc_percent_with_taxes(diesel_price, fuel_cost)

        records.append(
            {
                "Date": row.date,
                "Country": country,
                "Diesel price": round_numeric(diesel_price, 4),
                "FSC value, %": fsc_value,
            }
        )

    if not records:
        raise ValueError(
            "No rows matched countries from the calculation basis file."
        )

    result = pd.DataFrame(records, columns=list(OUTPUT_COLUMNS))
    result.sort_values(["Date", "Country"], inplace=True)
    result.reset_index(drop=True, inplace=True)
    return result


def format_fsc_worksheet(worksheet: Worksheet) -> None:
    """Apply table formatting to one FSC result worksheet."""
    if worksheet.max_row < 1:
        return

    headers = {
        worksheet.cell(1, column_index).value: column_index
        for column_index in range(1, worksheet.max_column + 1)
    }

    for column_name, width in COLUMN_WIDTHS.items():
        column_index = headers.get(column_name)
        if column_index is not None:
            worksheet.column_dimensions[
                worksheet.cell(1, column_index).column_letter
            ].width = width

    for row in worksheet.iter_rows(
        min_row=1,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=worksheet.max_column,
    ):
        for cell in row:
            cell.border = THIN_BORDER

            if cell.row == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")
                continue

            header = worksheet.cell(1, cell.column).value
            if header == "Date":
                cell.number_format = "DD.MM.YYYY"
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif header == "Country":
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif header == "Diesel price":
                cell.number_format = "0.0000"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif header == "FSC value, %":
                cell.number_format = '0.00"%"'
                cell.alignment = Alignment(horizontal="right", vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def format_fsc_workbook(workbook: Workbook) -> None:
    """Apply formatting to all sheets in the FSC result workbook."""
    for worksheet in workbook.worksheets:
        format_fsc_worksheet(worksheet)


def save_fsc_result_file(
    result_sheets: dict[str, pd.DataFrame],
    output_path: Path,
) -> None:
    """Save FSC result sheets to xlsx with formatting."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, result_df in result_sheets.items():
            result_df.to_excel(writer, sheet_name=sheet_name, index=False)
        format_fsc_workbook(writer.book)


def load_calculation_basis_sheets(
    basis_path: Path | None = None,
) -> tuple[str, str, dict[str, float], dict[str, float]]:
    """Load fuel costs for both calculation basis tabs."""
    basis_file = basis_path or find_calculation_basis_file()

    with pd.ExcelFile(basis_file, engine="openpyxl") as workbook:
        sheet_names = workbook.sheet_names

    if ROTALIS_PT_SHEET not in sheet_names:
        raise ValueError(
            f"Calculation basis file must contain a '{ROTALIS_PT_SHEET}' tab."
        )

    with_taxes_sheet = get_non_rotalis_sheet(sheet_names)
    fuel_costs_with_taxes = load_fuel_costs(basis_file, with_taxes_sheet)
    fuel_costs_wo_taxes = load_fuel_costs(basis_file, ROTALIS_PT_SHEET)

    return (
        with_taxes_sheet,
        ROTALIS_PT_SHEET,
        fuel_costs_with_taxes,
        fuel_costs_wo_taxes,
    )


def calculate_fsc_from_processed_file(
    processed_path: Path,
    basis_path: Path | None = None,
    output_dir: Path | None = None,
) -> tuple[dict[str, pd.DataFrame], Path]:
    """Calculate FSC values from a processed prices file."""
    with pd.ExcelFile(processed_path, engine="openpyxl") as workbook:
        sheet_names = workbook.sheet_names

    for required_sheet in (PROCESSING_SHEET_WITH_TAXES, PROCESSING_SHEET_WO_TAXES):
        if required_sheet not in sheet_names:
            raise ValueError(
                f"Processed file must contain a '{required_sheet}' tab."
            )

    (
        _with_taxes_basis_sheet,
        _wo_taxes_basis_sheet,
        fuel_costs_with_taxes,
        fuel_costs_wo_taxes,
    ) = load_calculation_basis_sheets(basis_path)

    prices_with_taxes = pd.read_excel(
        processed_path,
        sheet_name=PROCESSING_SHEET_WITH_TAXES,
        engine="openpyxl",
    )
    prices_wo_taxes = pd.read_excel(
        processed_path,
        sheet_name=PROCESSING_SHEET_WO_TAXES,
        engine="openpyxl",
    )

    result_sheets = {
        PROCESSING_SHEET_WITH_TAXES: build_fsc_dataframe(
            prices_with_taxes,
            fuel_costs_with_taxes,
            wo_taxes=False,
        ),
        PROCESSING_SHEET_WO_TAXES: build_fsc_dataframe(
            prices_wo_taxes,
            fuel_costs_wo_taxes,
            wo_taxes=True,
        ),
    }

    folder = output_dir or OUTPUT_DIR
    folder.mkdir(parents=True, exist_ok=True)
    output_path = folder / f"{processed_path.stem}_fsc.xlsx"
    save_fsc_result_file(result_sheets, output_path)

    return result_sheets, output_path


def run_calculator(
    processed_path: Path | None = None,
    basis_path: Path | None = None,
    output_dir: Path | None = None,
) -> tuple[dict[str, pd.DataFrame], Path]:
    """Run interactive or scripted FSC calculation."""
    selected_file = processed_path or choose_processed_file()

    print(f"\nUsing calculation basis: {(basis_path or find_calculation_basis_file()).name}")
    result_sheets, output_path = calculate_fsc_from_processed_file(
        processed_path=selected_file,
        basis_path=basis_path,
        output_dir=output_dir,
    )

    for sheet_name, result_df in result_sheets.items():
        print(f"  {sheet_name}: {len(result_df)} rows")

    print(f"Saved to: {output_path}")
    return result_sheets, output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Calculate FSC values from processed price files.",
    )
    parser.add_argument(
        "--processed",
        type=Path,
        help="Path to a processed xlsx file in the processing folder.",
    )
    parser.add_argument(
        "--basis",
        type=Path,
        help="Path to Calculation basis.xlsx (auto-detected if omitted).",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        help="Folder where result xlsx files will be saved (default: output).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    run_calculator(
        processed_path=args.processed,
        basis_path=args.basis,
        output_dir=args.output_dir,
    )


if __name__ == "__main__":
    main()

from __future__ import annotations

import argparse
import re
import shutil
from calendar import monthrange
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from config import ORIGINAL_FILE_DIR, OUTPUT_DIR

RATE_CARD_FILENAME = "Rate card export.xlsx"
FSC_SHEET_WITH_TAXES = "Prices with taxes"
FSC_SHEET_WO_TAXES = "Prices wo taxes"
ROTALIS_PT_MARKER = "rotalis pt"
FUEL_SURCHARGE_MARKER = "fuel surcharge"
EXPECTED_RATE_CARD_TABS = 5

COL_LANE = 1
COL_PERCENT = 2
COL_APPLIES = 3
COL_VALID_FROM = 4
COL_VALID_TO = 5

ORIGIN_COUNTRY_PATTERN = re.compile(
    r"Origin Country equals ((?:'[^']+'(?:,\s*)?)+)"
)
DESTINATION_COUNTRY_PATTERN = re.compile(
    r"Destination Country equals ((?:'[^']+'(?:,\s*)?)+)"
)
DEDICATED_SEA_PATTERN = re.compile(
    r"^TRANSPORT_MODE equals 'SEA' in any item\s*$",
    re.IGNORECASE,
)
FSC_OUTPUT_NAME_PATTERN = re.compile(r"^(?P<name>.+)_prices_\d{2}_\d{4}_fsc$", re.IGNORECASE)

PERIOD_FIRST = "first"
PERIOD_SECOND = "second"
PERIOD_BOTH = "both"


def choose_rate_card_period() -> set[str]:
    """Ask which half-month period(s) to include in the rate card result."""
    print("\nWhich rate card period do you want in the result?")
    print("  1. 01 to 15")
    print("  2. 16 to last day of month")
    print("  3. Both")

    choices = {
        "1": {PERIOD_FIRST},
        "2": {PERIOD_SECOND},
        "3": {PERIOD_FIRST, PERIOD_SECOND},
    }

    while True:
        choice = input("Enter choice (1, 2, or 3): ").strip()
        if choice in choices:
            return choices[choice]
        print("Please enter 1, 2, or 3.")


def parse_rate_card_period(value: str) -> set[str]:
    normalized = value.strip().casefold().replace("_", "-")
    mapping = {
        "first": {PERIOD_FIRST},
        "1-15": {PERIOD_FIRST},
        "01-15": {PERIOD_FIRST},
        PERIOD_SECOND: {PERIOD_SECOND},
        "16-end": {PERIOD_SECOND},
        "16-last": {PERIOD_SECOND},
        PERIOD_BOTH: {PERIOD_FIRST, PERIOD_SECOND},
        "all": {PERIOD_FIRST, PERIOD_SECOND},
    }
    if normalized not in mapping:
        raise ValueError(
            "Period must be one of: first, second, both, 1-15, 16-end"
        )
    return mapping[normalized]


def resolve_lane_target_half(
    *,
    valid_from: date | None,
    lane_row_count: int,
    available_halves: list[str],
    period_halves: set[str],
) -> str | None:
    """Pick which FSC half applies to one lane row for the selected period(s)."""
    selected_available = [half for half in available_halves if half in period_halves]
    if not selected_available:
        return None

    if valid_from:
        row_half = period_half_from_valid_from(valid_from)
        if row_half in period_halves and row_half in available_halves:
            return row_half

    if lane_row_count == 1 and len(selected_available) == 1:
        return selected_available[0]

    return None


def should_duplicate_second_half_row(
    *,
    lane_row_count: int,
    available_halves: list[str],
    period_halves: set[str],
) -> bool:
    if lane_row_count != 1:
        return False

    if PERIOD_FIRST not in period_halves or PERIOD_SECOND not in period_halves:
        return False

    return PERIOD_FIRST in available_halves and PERIOD_SECOND in available_halves


def list_fsc_output_files(output_dir: Path | None = None) -> list[Path]:
    folder = output_dir or OUTPUT_DIR
    if not folder.exists():
        raise FileNotFoundError(f"Output folder not found: {folder}")

    files = sorted(
        path for path in folder.glob("*_fsc.xlsx")
        if not path.name.startswith("~$")
    )
    if not files:
        raise FileNotFoundError(f"No FSC output files found in: {folder}")

    return files


def choose_fsc_output_file(output_dir: Path | None = None) -> Path:
    files = list_fsc_output_files(output_dir)

    print("Available FSC output files:")
    for index, file_path in enumerate(files, start=1):
        print(f"  {index}. {file_path.name}")

    while True:
        choice = input("Enter file number: ").strip()
        if not choice.isdigit():
            print("Please enter a valid number.")
            continue

        selected_index = int(choice)
        if 1 <= selected_index <= len(files):
            return files[selected_index - 1]

        print(f"Please enter a number between 1 and {len(files)}.")


def find_rate_card_export(original_dir: Path | None = None) -> Path:
    folder = original_dir or ORIGINAL_FILE_DIR
    direct_path = folder / RATE_CARD_FILENAME
    if direct_path.exists():
        return direct_path

    matches = sorted(folder.rglob(RATE_CARD_FILENAME))
    matches = [path for path in matches if not path.name.startswith("~$")]
    if not matches:
        raise FileNotFoundError(f"{RATE_CARD_FILENAME} was not found under input/rules.")

    return matches[0]


def parse_quoted_values(text: str) -> list[str]:
    return re.findall(r"'([^']+)'", text)


def extract_country_from_applies(applies: str | None) -> str | None:
    if not applies:
        return None

    text = str(applies)
    origin_match = ORIGIN_COUNTRY_PATTERN.search(text)
    destination_match = DESTINATION_COUNTRY_PATTERN.search(text)

    origin_values = parse_quoted_values(origin_match.group(1)) if origin_match else []
    destination_values = (
        parse_quoted_values(destination_match.group(1)) if destination_match else []
    )

    if origin_match and destination_match:
        if len(destination_values) == 1:
            return destination_values[0].upper()
        if len(origin_values) == 1:
            return origin_values[0].upper()
        return destination_values[0].upper() if destination_values else None

    if destination_match:
        return destination_values[0].upper() if destination_values else None

    if origin_match:
        if len(origin_values) == 1:
            return origin_values[0].upper()
        return origin_values[0].upper() if origin_values else None

    return None


def is_sea_lane(applies: str | None) -> bool:
    if not applies:
        return False

    text = str(applies).strip()
    if "does not equal" in text.casefold():
        return False

    return bool(DEDICATED_SEA_PATTERN.match(text))


def sheet_uses_wo_taxes(sheet_name: str) -> bool:
    return ROTALIS_PT_MARKER in sheet_name.casefold()


def is_fuel_surcharge_sheet(sheet_name: str) -> bool:
    normalized = sheet_name.casefold().replace("-", " ").replace("_", " ")
    return FUEL_SURCHARGE_MARKER in normalized


def get_rate_card_worksheets(workbook) -> list[Worksheet]:
    """Return all Fuel Surcharge tabs that should be updated."""
    worksheets = [
        worksheet
        for worksheet in workbook.worksheets
        if is_fuel_surcharge_sheet(worksheet.title)
    ]

    if len(worksheets) < EXPECTED_RATE_CARD_TABS:
        found = [worksheet.title for worksheet in worksheets]
        raise ValueError(
            "Rate card export must contain 5 Fuel Surcharge tabs "
            f"({EXPECTED_RATE_CARD_TABS} expected, found {len(worksheets)}): {found}"
        )

    return worksheets


def describe_fsc_source(sheet_name: str) -> str:
    if sheet_uses_wo_taxes(sheet_name):
        return FSC_SHEET_WO_TAXES
    return FSC_SHEET_WITH_TAXES


def choose_fsc_lookup_for_sheet(
    sheet_name: str,
    lookup_with_taxes: dict[tuple[str, str], float],
    lookup_wo_taxes: dict[tuple[str, str], float],
) -> dict[tuple[str, str], float]:
    if sheet_uses_wo_taxes(sheet_name):
        return lookup_wo_taxes
    return lookup_with_taxes


def parse_excel_date(value: object) -> date | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%d.%m.%Y", "%d.%m.%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    if pd.notna(parsed):
        return parsed.date()

    return None


def format_excel_date(value: date) -> str:
    return value.strftime("%d.%m.%Y")


def format_percent_value(value: float) -> str:
    """Format percentage values with a dot decimal separator."""
    return f"{round(value, 2):.2f}"


def derive_rate_card_output_name(
    fsc_output_path: Path,
    fsc_input_path: Path | None = None,
) -> str:
    """Derive the output filename from the original FSC input file."""
    if fsc_input_path is not None:
        stem = fsc_input_path.stem
    else:
        match = FSC_OUTPUT_NAME_PATTERN.match(fsc_output_path.stem)
        if match:
            stem = match.group("name")
        elif fsc_output_path.stem.endswith("_fsc"):
            stem = fsc_output_path.stem[:-4]
        else:
            stem = fsc_output_path.stem

    return f"{stem} result.xlsx"


def period_half_for_date(value: date) -> str:
    distance_to_first = abs(value.day - 1)
    distance_to_sixteenth = abs(value.day - 16)
    return "first" if distance_to_first <= distance_to_sixteenth else "second"


def period_bounds(year: int, month: int, half: str) -> tuple[date, date]:
    if half == "first":
        return date(year, month, 1), date(year, month, 15)

    last_day = monthrange(year, month)[1]
    return date(year, month, 16), date(year, month, last_day)


def period_half_from_valid_from(valid_from: date) -> str:
    return "first" if valid_from.day <= 15 else "second"


def build_fsc_lookup(fsc_df: pd.DataFrame) -> tuple[dict[tuple[str, str], float], int, int]:
    lookup: dict[tuple[str, str], float] = {}
    months: set[tuple[int, int]] = set()

    for _, row in fsc_df.iterrows():
        row_date = parse_excel_date(row["Date"])
        if row_date is None:
            continue

        country = str(row["Country"]).strip().upper()
        half = period_half_for_date(row_date)
        lookup[(country, half)] = float(row["FSC value, %"])
        months.add((row_date.month, row_date.year))

    if not lookup or not months:
        raise ValueError("FSC output file does not contain any usable FSC values.")

    if len(months) > 1:
        raise ValueError("FSC output file contains multiple months. Process one month at a time.")

    month, year = next(iter(months))
    return lookup, month, year


def load_fsc_lookups(
    fsc_output_path: Path,
) -> tuple[dict[tuple[str, str], float], dict[tuple[str, str], float], int, int]:
    with pd.ExcelFile(fsc_output_path, engine="openpyxl") as workbook:
        sheet_names = workbook.sheet_names

    for required_sheet in (FSC_SHEET_WITH_TAXES, FSC_SHEET_WO_TAXES):
        if required_sheet not in sheet_names:
            raise ValueError(f"FSC output file must contain a '{required_sheet}' tab.")

    with_taxes_df = pd.read_excel(
        fsc_output_path,
        sheet_name=FSC_SHEET_WITH_TAXES,
        engine="openpyxl",
    )
    wo_taxes_df = pd.read_excel(
        fsc_output_path,
        sheet_name=FSC_SHEET_WO_TAXES,
        engine="openpyxl",
    )

    lookup_with_taxes, month, year = build_fsc_lookup(with_taxes_df)
    lookup_wo_taxes, month_wo, year_wo = build_fsc_lookup(wo_taxes_df)

    if (month, year) != (month_wo, year_wo):
        raise ValueError("Both FSC output tabs must refer to the same month.")

    return lookup_with_taxes, lookup_wo_taxes, month, year


def get_cell_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def is_lane_header_row(worksheet: Worksheet, row_index: int) -> bool:
    applies_header = get_cell_text(worksheet.cell(row_index, COL_APPLIES).value)
    percent_header = get_cell_text(worksheet.cell(row_index, COL_PERCENT).value)
    if applies_header != "Applies if":
        return False

    if not percent_header:
        return False

    normalized_header = percent_header.casefold().replace("-", " ").replace("_", " ")
    return "over cost" in normalized_header


def find_lane_header_rows(worksheet: Worksheet) -> list[int]:
    header_rows = [
        row_index
        for row_index in range(1, worksheet.max_row + 1)
        if is_lane_header_row(worksheet, row_index)
    ]
    if not header_rows:
        raise ValueError("Could not find lane header rows in the rate card export.")
    return header_rows


def is_lane_data_row(worksheet: Worksheet, row_index: int) -> bool:
    applies = get_cell_text(worksheet.cell(row_index, COL_APPLIES).value)
    valid_from = worksheet.cell(row_index, COL_VALID_FROM).value
    return bool(applies and valid_from is not None and str(valid_from).strip())


def update_lane_row(
    worksheet: Worksheet,
    row_index: int,
    *,
    percent_value: float,
    valid_from: date,
    valid_to: date,
) -> None:
    worksheet.cell(row_index, COL_PERCENT, format_percent_value(percent_value))
    worksheet.cell(row_index, COL_VALID_FROM, format_excel_date(valid_from))
    worksheet.cell(row_index, COL_VALID_TO, format_excel_date(valid_to))


def duplicate_lane_row(worksheet: Worksheet, source_row: int, target_row: int) -> None:
    for column_index in range(COL_LANE, COL_VALID_TO + 1):
        worksheet.cell(target_row, column_index).value = worksheet.cell(
            source_row, column_index
        ).value


def update_worksheet_lanes(
    worksheet: Worksheet,
    *,
    fsc_lookup: dict[tuple[str, str], float],
    month: int,
    year: int,
    period_halves: set[str] | None = None,
) -> int:
    """Update all lane sections in one worksheet."""
    if period_halves is None:
        period_halves = {PERIOD_FIRST, PERIOD_SECOND}

    header_rows = find_lane_header_rows(worksheet)
    updated_rows = 0

    for section_index, header_row in enumerate(header_rows):
        next_header = (
            header_rows[section_index + 1]
            if section_index + 1 < len(header_rows)
            else worksheet.max_row + 1
        )

        row_indexes = [
            row_index
            for row_index in range(header_row + 1, next_header)
            if is_lane_data_row(worksheet, row_index)
        ]

        for row_index in row_indexes:
            applies = get_cell_text(worksheet.cell(row_index, COL_APPLIES).value)
            if not applies:
                continue

            if is_sea_lane(applies):
                sea_rows = [
                    lane_row
                    for lane_row in row_indexes
                    if is_sea_lane(
                        get_cell_text(worksheet.cell(lane_row, COL_APPLIES).value)
                    )
                ]
                if row_index != sea_rows[0]:
                    continue

                lane_rows = list(sea_rows)
                if should_duplicate_second_half_row(
                    lane_row_count=len(lane_rows),
                    available_halves=[PERIOD_FIRST, PERIOD_SECOND],
                    period_halves=period_halves,
                ):
                    new_row = next_header
                    worksheet.insert_rows(new_row)
                    duplicate_lane_row(worksheet, lane_rows[0], new_row)
                    second_from, second_to = period_bounds(year, month, PERIOD_SECOND)
                    worksheet.cell(new_row, COL_VALID_FROM, format_excel_date(second_from))
                    worksheet.cell(new_row, COL_VALID_TO, format_excel_date(second_to))
                    lane_rows.append(new_row)
                    next_header += 1

                for lane_row in lane_rows:
                    valid_from = parse_excel_date(
                        worksheet.cell(lane_row, COL_VALID_FROM).value
                    )
                    half = resolve_lane_target_half(
                        valid_from=valid_from,
                        lane_row_count=len(lane_rows),
                        available_halves=[PERIOD_FIRST, PERIOD_SECOND],
                        period_halves=period_halves,
                    )
                    if half is None:
                        continue

                    period_from, period_to = period_bounds(year, month, half)
                    update_lane_row(
                        worksheet,
                        lane_row,
                        percent_value=0,
                        valid_from=period_from,
                        valid_to=period_to,
                    )
                    updated_rows += 1
                continue

            country = extract_country_from_applies(applies)
            if not country:
                continue

            lane_rows = [row_index]
            available_halves = [
                half
                for half in (PERIOD_FIRST, PERIOD_SECOND)
                if (country, half) in fsc_lookup
            ]
            if not available_halves:
                continue

            if should_duplicate_second_half_row(
                lane_row_count=len(lane_rows),
                available_halves=available_halves,
                period_halves=period_halves,
            ):
                new_row = next_header
                worksheet.insert_rows(new_row)
                duplicate_lane_row(worksheet, lane_rows[0], new_row)
                second_from, second_to = period_bounds(year, month, PERIOD_SECOND)
                worksheet.cell(new_row, COL_VALID_FROM, format_excel_date(second_from))
                worksheet.cell(new_row, COL_VALID_TO, format_excel_date(second_to))
                lane_rows.append(new_row)
                next_header += 1

            for lane_row in lane_rows:
                valid_from = parse_excel_date(
                    worksheet.cell(lane_row, COL_VALID_FROM).value
                )
                half = resolve_lane_target_half(
                    valid_from=valid_from,
                    lane_row_count=len(lane_rows),
                    available_halves=available_halves,
                    period_halves=period_halves,
                )
                if half is None or (country, half) not in fsc_lookup:
                    continue

                period_from, period_to = period_bounds(year, month, half)
                update_lane_row(
                    worksheet,
                    lane_row,
                    percent_value=fsc_lookup[(country, half)],
                    valid_from=period_from,
                    valid_to=period_to,
                )
                updated_rows += 1

    return updated_rows


def update_rate_card_export(
    fsc_output_path: Path,
    rate_card_path: Path | None = None,
    fsc_input_path: Path | None = None,
    output_dir: Path | None = None,
    period_halves: set[str] | None = None,
) -> tuple[Path, int, dict[str, int]]:
    """Update % over cost values in a copy of the rate card export."""
    source_rate_card = rate_card_path or find_rate_card_export()
    lookup_with_taxes, lookup_wo_taxes, month, year = load_fsc_lookups(fsc_output_path)

    folder = output_dir or OUTPUT_DIR
    folder.mkdir(parents=True, exist_ok=True)
    output_name = derive_rate_card_output_name(fsc_output_path, fsc_input_path)
    output_path = folder / output_name
    shutil.copy2(source_rate_card, output_path)

    workbook = load_workbook(output_path)
    updated_rows = 0
    updates_by_sheet: dict[str, int] = {}

    for worksheet in get_rate_card_worksheets(workbook):
        fsc_lookup = choose_fsc_lookup_for_sheet(
            worksheet.title,
            lookup_with_taxes,
            lookup_wo_taxes,
        )
        sheet_updates = update_worksheet_lanes(
            worksheet,
            fsc_lookup=fsc_lookup,
            month=month,
            year=year,
            period_halves=period_halves,
        )
        updates_by_sheet[worksheet.title] = sheet_updates
        updated_rows += sheet_updates
        print(
            f"  {worksheet.title}: {sheet_updates} rows "
            f"({describe_fsc_source(worksheet.title)})"
        )

    workbook.save(output_path)
    return output_path, updated_rows, updates_by_sheet


def run_rate_card_updater(
    fsc_output_path: Path | None = None,
    rate_card_path: Path | None = None,
    fsc_input_path: Path | None = None,
    output_dir: Path | None = None,
    period_halves: set[str] | None = None,
) -> tuple[Path, int, dict[str, int]]:
    selected_fsc_output = fsc_output_path or choose_fsc_output_file()
    source_rate_card = rate_card_path or find_rate_card_export()
    selected_periods = period_halves or choose_rate_card_period()

    print(f"\nUsing FSC output:     {selected_fsc_output.name}")
    print(f"Using rate card:      {source_rate_card.name}")
    if selected_periods == {PERIOD_FIRST}:
        print("Rate card period:     01 to 15")
    elif selected_periods == {PERIOD_SECOND}:
        print("Rate card period:     16 to last day of month")
    else:
        print("Rate card period:     Both halves")

    output_path, updated_rows, updates_by_sheet = update_rate_card_export(
        fsc_output_path=selected_fsc_output,
        rate_card_path=source_rate_card,
        fsc_input_path=fsc_input_path,
        output_dir=output_dir,
        period_halves=selected_periods,
    )

    print(f"Updated rows:         {updated_rows}")
    print(f"Saved to:             {output_path}")
    return output_path, updated_rows, updates_by_sheet


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Update rate card export % over cost values from FSC output.",
    )
    parser.add_argument(
        "--fsc-output",
        type=Path,
        help="Path to an FSC output xlsx file.",
    )
    parser.add_argument(
        "--fsc-input",
        type=Path,
        help="Path to the original FSC input file (used for the output filename).",
    )
    parser.add_argument(
        "--rate-card",
        type=Path,
        help="Path to the original Rate card export.xlsx file.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        help="Folder where the updated rate card will be saved.",
    )
    parser.add_argument(
        "--period",
        choices=("first", "second", "both"),
        help="Rate card period to update: first (01-15), second (16-end), or both.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    period_halves = parse_rate_card_period(args.period) if args.period else None
    run_rate_card_updater(
        fsc_output_path=args.fsc_output,
        rate_card_path=args.rate_card,
        fsc_input_path=args.fsc_input,
        output_dir=args.output_dir,
        period_halves=period_halves,
    )


if __name__ == "__main__":
    main()
